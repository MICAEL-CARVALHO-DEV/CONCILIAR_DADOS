from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Callable

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..models import Emenda, EmendaLock, ExportLog, Historico, ImportGovernancaLog, ImportLinha, ImportLote, SupportThread


IMPORT_ALIASES = {
    "id": ["id", "id_interno", "id interno", "codigo_interno", "codigo interno", "emenda"],
    "ano": ["ano", "exercicio"],
    "identificacao": ["identificacao", "identificacao_emenda", "numero_emenda", "identificacao da emenda"],
    "cod_subfonte": ["cod_subfonte", "codigo_subfonte", "subfonte", "cod subfonte", "cod. subfonte"],
    "deputado": ["deputado", "autor", "parlamentar"],
    "cod_uo": ["cod_uo", "codigo_uo", "uo", "cod uo", "cod. uo"],
    "sigla_uo": ["sigla_uo", "sigla uo", "uo_sigla", "sigla da uo", "sigla do uo", "sigla do orgao", "sigla do órgão"],
    "cod_orgao": ["cod_orgao", "codigo_orgao", "orgao", "cod orgao", "cod. orgao", "cod. órgão"],
    "cod_acao": ["cod_acao", "codigo_acao", "acao", "cod acao", "cod da acao", "cod. da acao", "codigo da acao", "cód. da ação"],
    "descricao_acao": ["descricao_acao", "descricao da acao", "acao_descricao", "descricao", "descritor da acao", "descritor da ação"],
    "plan_a": ["plan_a", "plano_a", "plano a", "planoa", "plano a acao", "plano de acao a"],
    "plan_b": ["plan_b", "plano_b", "plano b", "planob", "plano b acao", "plano de acao b"],
    "municipio": ["municipio", "cidade", "municipio / estado", "municipio estado", "município / estado"],
    "valor_inicial": ["valor_inicial", "valor inicial", "valor_original", "valor original", "valor inicial epi"],
    "valor_atual": ["valor_atual", "valor atual", "valor", "valor_emenda", "valor emenda", "valor atual epi"],
    "processo_sei": ["processo_sei", "processo sei", "sei", "processo"],
    "status_oficial": ["status_oficial", "status oficial", "status"],
}

REFERENCE_FIELDS = ["identificacao", "cod_subfonte", "cod_acao", "municipio", "deputado"]
TRACKED_FIELDS = [
    ("ano", "number"),
    ("identificacao", "string"),
    ("cod_subfonte", "string"),
    ("deputado", "string"),
    ("cod_uo", "string"),
    ("sigla_uo", "string"),
    ("cod_orgao", "string"),
    ("cod_acao", "string"),
    ("descricao_acao", "string"),
    ("plan_a", "string"),
    ("plan_b", "string"),
    ("municipio", "string"),
    ("valor_inicial", "money"),
    ("valor_atual", "money"),
    ("processo_sei", "string"),
]
STATUS_VALUES = [
    "Recebido",
    "Em analise",
    "Pendente",
    "Aguardando execucao",
    "Em execucao",
    "Aprovado",
    "Concluido",
    "Cancelado",
]
HEADER_HINTS = {
    "identificacao",
    "deputado",
    "status",
    "municipio",
    "cod_uo",
    "cod_subfonte",
    "cod_da_acao",
    "descritor_da_acao",
}


def _as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_loose_text(value) -> str:
    return (
        unicodedata.normalize("NFD", _as_text(value))
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
        .strip()
    )


def _normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _normalize_loose_text(value)).strip("_")


def _normalize_reference_part(value) -> str:
    return re.sub(r"\s+", " ", _normalize_loose_text(value)).strip()


def _build_reference_key(record: dict) -> str:
    parts = [_normalize_reference_part(record.get(field)) for field in REFERENCE_FIELDS]
    if all(not part for part in parts):
        return ""
    return "|".join(parts)


def _to_int(value) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return 0


def _to_number(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    raw = _as_text(value).replace(" ", "")
    raw = re.sub(r"\.(?=\d{3}(\D|$))", "", raw)
    raw = raw.replace(",", ".")
    raw = re.sub(r"[^\d.\-]", "", raw)
    try:
        return float(raw)
    except Exception:
        return 0.0


def _to_number_or_none(value) -> float | None:
    if value is None:
        return None
    if _as_text(value) == "":
        return None
    return _to_number(value)


def _normalize_row_keys(row: dict) -> dict:
    out: dict[str, object] = {}
    for key, value in (row or {}).items():
        normalized = _normalize_header(key)
        if not normalized:
            continue
        if normalized not in out or _as_text(out.get(normalized)) == "":
            out[normalized] = value
    return out


def _pick_value(normalized_row: dict, aliases: list[str]) -> object:
    for alias in aliases or []:
        normalized = _normalize_header(alias)
        raw = normalized_row.get(normalized)
        if raw is None:
            continue
        if _as_text(raw) != "":
            return raw
    return ""


def _normalize_status(value) -> str:
    cleaned = _normalize_loose_text(value)
    if not cleaned:
        return "Recebido"
    for status in STATUS_VALUES:
        if _normalize_loose_text(status) == cleaned:
            return status
    return "Recebido"


def _build_headers_from_row(raw_header: list[object]) -> list[str]:
    out: list[str] = []
    used: dict[str, int] = {}
    header_row = raw_header or []
    total = max(len(header_row), 1)
    for idx in range(total):
        base = _as_text(header_row[idx]) or f"COL_{idx + 1}"
        key = base
        suffix = 2
        while key in used:
            key = f"{base}_{suffix}"
            suffix += 1
        used[key] = 1
        out.append(key)
    return out


def _is_row_empty(arr: list[object]) -> bool:
    return not any(_as_text(value) for value in (arr or []))


def _row_array_to_object(arr: list[object], headers: list[str]) -> dict[str, str]:
    obj: dict[str, str] = {}
    row = arr or []
    for idx, key in enumerate(headers or []):
        if not key:
            continue
        obj[key] = _as_text(row[idx] if idx < len(row) else "")
    return obj


def _build_known_header_set() -> set[str]:
    headers = set()
    for alias_list in IMPORT_ALIASES.values():
        for alias in alias_list:
            headers.add(_normalize_header(alias))
    return headers


def _detect_header_row(matrix: list[list[object]]) -> dict | None:
    if not matrix:
        return None

    best_index = -1
    best_score = -1
    scan_limit = min(len(matrix), 40)
    for idx in range(scan_limit):
        row = matrix[idx] or []
        non_empty = sum(1 for value in row if _as_text(value))
        if non_empty < 3:
            continue
        normalized = [_normalize_header(value) for value in row]
        score = non_empty
        for hint in HEADER_HINTS:
            if hint in normalized:
                score += 5
        if score > best_score:
            best_score = score
            best_index = idx

    if best_index < 0:
        return None
    return {"headerRowIndex": best_index, "headers": _build_headers_from_row(matrix[best_index] or [])}


def _read_sheet_matrix(worksheet) -> list[list[object]]:
    matrix: list[list[object]] = []
    for row in worksheet.iter_rows(values_only=True):
        matrix.append(list(row or []))
    return matrix


def _extract_planilha1_aoa_from_workbook(workbook) -> list[list[str]] | None:
    ws = workbook["Planilha1"] if "Planilha1" in workbook.sheetnames else None
    if ws is None:
        return None

    matrix = _read_sheet_matrix(ws)
    if not matrix:
        return None

    header_idx = -1
    scan_limit = min(len(matrix), 50)
    for idx in range(scan_limit):
        row = matrix[idx] or []
        c1 = _normalize_header(row[0] if len(row) > 0 else "")
        c2 = _normalize_header(row[1] if len(row) > 1 else "")
        has_rotulo = "rotulos_de_linha" in c1 or "rotulo_de_linha" in c1
        has_contagem = "contagem_de_deputado" in c2
        if has_rotulo and has_contagem:
            header_idx = idx
            break

    if header_idx < 0:
        return None

    header_row = matrix[header_idx] or []
    out = [[_as_text(header_row[0]) or "Rotulos de Linha", _as_text(header_row[1] if len(header_row) > 1 else "") or "Contagem de Deputado"]]
    for idx in range(header_idx + 1, len(matrix)):
        row = matrix[idx] or []
        label = _as_text(row[0] if len(row) > 0 else "")
        value = _as_text(row[1] if len(row) > 1 else "")
        if not label and not value:
            continue
        out.append([label, value])
        if _normalize_loose_text(label) == "total geral":
            break
    return out if len(out) > 1 else None


def _map_import_row(ctx: dict) -> dict:
    row = _normalize_row_keys((ctx or {}).get("row") or {})
    ano = _to_int(_pick_value(row, IMPORT_ALIASES["ano"]))
    row_id = _as_text(_pick_value(row, IMPORT_ALIASES["id"]))
    record = {
        "id": row_id,
        "ano": ano or datetime.utcnow().year,
        "identificacao": _as_text(_pick_value(row, IMPORT_ALIASES["identificacao"])) or row_id,
        "cod_subfonte": _as_text(_pick_value(row, IMPORT_ALIASES["cod_subfonte"])),
        "deputado": _as_text(_pick_value(row, IMPORT_ALIASES["deputado"])),
        "cod_uo": _as_text(_pick_value(row, IMPORT_ALIASES["cod_uo"])),
        "sigla_uo": _as_text(_pick_value(row, IMPORT_ALIASES["sigla_uo"])),
        "cod_orgao": _as_text(_pick_value(row, IMPORT_ALIASES["cod_orgao"])),
        "cod_acao": _as_text(_pick_value(row, IMPORT_ALIASES["cod_acao"])),
        "descricao_acao": _as_text(_pick_value(row, IMPORT_ALIASES["descricao_acao"])),
        "plan_a": _as_text(_pick_value(row, IMPORT_ALIASES["plan_a"])),
        "plan_b": _as_text(_pick_value(row, IMPORT_ALIASES["plan_b"])),
        "municipio": _as_text(_pick_value(row, IMPORT_ALIASES["municipio"])),
        "valor_inicial": _to_number_or_none(_pick_value(row, IMPORT_ALIASES["valor_inicial"])),
        "valor_atual": _to_number_or_none(_pick_value(row, IMPORT_ALIASES["valor_atual"])),
        "processo_sei": _as_text(_pick_value(row, IMPORT_ALIASES["processo_sei"])),
        "status_oficial": "",
        "all_fields": dict((ctx or {}).get("row") or {}),
        "source_sheet": (ctx or {}).get("sheetName") or "XLSX",
        "source_row": (ctx or {}).get("rowNumber"),
    }
    status_raw = _as_text(_pick_value(row, IMPORT_ALIASES["status_oficial"]))
    if status_raw:
        record["status_oficial"] = _normalize_status(status_raw)
    record["ref_key"] = _build_reference_key(record)
    return record


def _serialize_preview_source_row(ctx: dict, incoming: dict, status_linha: str, mensagem: str) -> dict:
    source = ctx or {}
    mapped = incoming or {}
    raw_row = dict(source.get("row") or {})
    normalized_status = str(status_linha or "").strip().upper()
    serialized_row = dict(raw_row)
    serialized_row.update(
        {
            "id": _as_text(mapped.get("id")),
            "ano": _as_text(mapped.get("ano")),
            "identificacao": _as_text(mapped.get("identificacao")),
            "cod_subfonte": _as_text(mapped.get("cod_subfonte")),
            "deputado": _as_text(mapped.get("deputado")),
            "cod_uo": _as_text(mapped.get("cod_uo")),
            "sigla_uo": _as_text(mapped.get("sigla_uo")),
            "cod_orgao": _as_text(mapped.get("cod_orgao")),
            "cod_acao": _as_text(mapped.get("cod_acao")),
            "descricao_acao": _as_text(mapped.get("descricao_acao")),
            "plan_a": _as_text(mapped.get("plan_a")),
            "plan_b": _as_text(mapped.get("plan_b")),
            "municipio": _as_text(mapped.get("municipio")),
            "valor_inicial": _as_text(mapped.get("valor_inicial")),
            "valor_atual": _as_text(mapped.get("valor_atual")),
            "processo_sei": _as_text(mapped.get("processo_sei")),
            "status_oficial": _as_text(mapped.get("status_oficial")),
            "ref_key": _as_text(mapped.get("ref_key")),
            "__previewStatus": normalized_status,
            "__previewMessage": _as_text(mensagem),
        }
    )
    return {
        "aba": str(source.get("sheetName") or "XLSX"),
        "linha": int(source.get("rowNumber") or 0),
        "dados": serialized_row,
    }


def _coerce_preview_source_rows(source_rows: list[dict], row_details: list[dict]) -> list[dict]:
    coerced: list[dict] = []
    details = row_details or []
    for index, item in enumerate(source_rows or []):
        source = item or {}
        row_data = dict(source.get("dados") or {})
        if row_data.get("__previewStatus"):
            coerced.append(source)
            continue
        detail = details[index] if index < len(details) else {}
        ctx = {
            "sheetName": source.get("aba") or source.get("sheetName") or "XLSX",
            "rowNumber": source.get("linha") if source.get("linha") is not None else source.get("rowNumber"),
            "row": row_data or source.get("row") or {},
        }
        incoming = _map_import_row(ctx)
        coerced.append(
            _serialize_preview_source_row(
                ctx,
                incoming,
                detail.get("status_linha") or "UNCHANGED",
                detail.get("mensagem") or "",
            )
        )
    return coerced


def _has_useful_data(record: dict) -> bool:
    checks = [
        record.get("id"),
        record.get("identificacao"),
        record.get("cod_subfonte"),
        record.get("cod_acao"),
        record.get("municipio"),
        record.get("deputado"),
        record.get("processo_sei"),
        record.get("ref_key"),
    ]
    has_text = any(_as_text(value) for value in checks)
    has_number = record.get("valor_inicial") is not None or record.get("valor_atual") is not None
    return has_text or has_number


def _has_incoming_value(value, field_type: str) -> bool:
    if field_type in {"money", "number"}:
        return value is not None and _as_text(value) != ""
    return _as_text(value) != ""


def _field_changed(previous, next_value, field_type: str) -> bool:
    if field_type == "money":
        return _to_number(previous) != _to_number(next_value)
    if field_type == "number":
        return _to_int(previous) != _to_int(next_value)
    return _normalize_loose_text(previous) != _normalize_loose_text(next_value)


def _build_import_validation_report(source_rows: list[dict]) -> dict:
    known_headers = _build_known_header_set()
    headers_found: list[str] = []
    header_seen: set[str] = set()
    preview_rows: list[dict] = []
    alerts: list[str] = []

    for ctx in source_rows[:5]:
        preview_rows.append(
            {
                "aba": ctx.get("sheetName") or "XLSX",
                "linha": int(ctx.get("rowNumber")) if ctx.get("rowNumber") is not None else None,
                "dados": dict(ctx.get("row") or {}),
            }
        )

    for ctx in source_rows:
        row = ctx.get("row") or {}
        for key in row.keys():
            if key not in header_seen:
                header_seen.add(key)
                headers_found.append(key)

    recognized = [header for header in headers_found if _normalize_header(header) in known_headers]
    unrecognized = [header for header in headers_found if _normalize_header(header) not in known_headers]
    header_normalized_count: dict[str, int] = {}
    for header in headers_found:
        normalized = _normalize_header(header)
        header_normalized_count[normalized] = header_normalized_count.get(normalized, 0) + 1
    duplicated = [key for key, count in header_normalized_count.items() if count > 1]

    if len(recognized) < 3:
        alerts.append("Cabecalho suspeito: poucas colunas reconhecidas.")
    if duplicated:
        alerts.append(f"Colunas duplicadas detectadas: {', '.join(duplicated)}")

    empties = {"identificacao": 0, "deputado": 0, "municipio": 0}
    for ctx in source_rows:
        mapped = _map_import_row(ctx)
        if not _as_text(mapped.get("identificacao")):
            empties["identificacao"] += 1
        if not _as_text(mapped.get("deputado")):
            empties["deputado"] += 1
        if not _as_text(mapped.get("municipio")):
            empties["municipio"] += 1

    for key, value in empties.items():
        if value > 0:
            alerts.append(f"Campo critico vazio em {key}: {value} linha(s).")

    sample = [_map_import_row(ctx) for ctx in source_rows[:50]]
    detected_types = {
        "ano": "numero" if any(item.get("ano") for item in sample) else "vazio",
        "valor_inicial": "numero" if any(item.get("valor_inicial") is not None for item in sample) else "vazio",
        "valor_atual": "numero" if any(item.get("valor_atual") is not None for item in sample) else "vazio",
        "identificacao": "texto" if any(_as_text(item.get("identificacao")) for item in sample) else "vazio",
        "processo_sei": "texto" if any(_as_text(item.get("processo_sei")) for item in sample) else "vazio",
    }

    return {
        "recognizedHeaders": recognized,
        "unrecognizedHeaders": unrecognized,
        "duplicatedHeaders": duplicated,
        "previewRows": preview_rows,
        "detectedTypes": detected_types,
        "alerts": alerts,
    }


def preview_import_xlsx_service(*, file_name: str, file_bytes: bytes, db: Session) -> dict:
    if not file_bytes:
        raise HTTPException(status_code=400, detail="arquivo vazio")
    if not str(file_name or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Formato nao suportado. Use apenas XLSX.")

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"falha ao ler arquivo XLSX: {exc}") from exc

    preferred_sheet = "Controle de EPI" if "Controle de EPI" in workbook.sheetnames else None
    ordered_sheet_names = [preferred_sheet] + [name for name in workbook.sheetnames if name != preferred_sheet] if preferred_sheet else list(workbook.sheetnames)
    known_headers = _build_known_header_set()
    source_rows: list[dict] = []
    read_sheets: list[str] = []

    for sheet_name in ordered_sheet_names:
        sheet = workbook[sheet_name]
        matrix = _read_sheet_matrix(sheet)
        detected = _detect_header_row(matrix)
        if not detected:
            continue

        headers = detected["headers"] or []
        recognized_count = sum(1 for header in headers if _normalize_header(header) in known_headers)

        if sheet_name == "Controle de EPI" and recognized_count < 5:
            raise HTTPException(
                status_code=400,
                detail="Cabecalho da aba Controle de EPI nao reconhecido. Verifique se a linha de cabecalho esta correta.",
            )
        if sheet_name != "Controle de EPI" and recognized_count < 3:
            continue

        read_sheets.append(sheet_name)
        for row_index in range(detected["headerRowIndex"] + 1, len(matrix)):
            arr = matrix[row_index] or []
            if _is_row_empty(arr):
                continue
            source_rows.append(
                {
                    "sheetName": sheet_name,
                    "rowNumber": row_index + 1,
                    "row": _row_array_to_object(arr, headers),
                }
            )

    if not source_rows:
        raise HTTPException(status_code=400, detail="nenhuma linha valida foi encontrada no arquivo")

    validation = _build_import_validation_report(source_rows)
    planilha1_aoa = _extract_planilha1_aoa_from_workbook(workbook)

    existing_records = db.query(Emenda).filter(Emenda.is_current.is_(True)).all()
    existing_by_id = {str(item.id_interno or "").strip(): item for item in existing_records if str(item.id_interno or "").strip()}
    existing_by_ref = {}
    for item in existing_records:
        ref_key = _build_reference_key(
            {
                "identificacao": item.identificacao,
                "cod_subfonte": item.cod_subfonte,
                "cod_acao": item.cod_acao,
                "municipio": item.municipio,
                "deputado": item.deputado,
            }
        )
        if ref_key and ref_key not in existing_by_ref:
            existing_by_ref[ref_key] = item

    row_details: list[dict] = []
    new_rows_preview: list[dict] = []
    seen_ids_in_file: set[str] = set()
    seen_refs_in_file: set[str] = set()
    normalized_source_rows: list[dict] = []
    report = {
        "fileName": file_name or "import.xlsx",
        "fileHash": hashlib.sha256(file_bytes).hexdigest(),
        "totalRows": len(source_rows),
        "consideredRows": 0,
        "skippedRows": 0,
        "invalidRows": 0,
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "duplicateById": 0,
        "duplicateByRef": 0,
        "duplicateInFile": 0,
        "conflictIdVsRef": 0,
        "sheetNames": read_sheets,
        "rowDetails": row_details,
        "newRowsPreview": new_rows_preview,
        "sourceRows": normalized_source_rows,
        "validation": validation,
        "planilha1Aoa": planilha1_aoa,
    }

    for ctx in source_rows:
        incoming = _map_import_row(ctx)
        ordem = len(row_details) + 1
        id_interno = incoming.get("id") or ""
        ref_key = incoming.get("ref_key") or ""

        if not _has_useful_data(incoming):
            report["invalidRows"] += 1
            mensagem = "Linha ignorada: sem dados uteis"
            row_details.append(
                {
                    "ordem": ordem,
                    "sheet_name": ctx.get("sheetName") or "XLSX",
                    "row_number": int(ctx.get("rowNumber") or 0),
                    "status_linha": "SKIPPED",
                    "id_interno": id_interno,
                    "ref_key": ref_key,
                    "mensagem": mensagem,
                }
            )
            normalized_source_rows.append(_serialize_preview_source_row(ctx, incoming, "SKIPPED", mensagem))
            continue

        report["consideredRows"] += 1

        duplicate_in_file = False
        if id_interno:
            if id_interno in seen_ids_in_file:
                duplicate_in_file = True
            seen_ids_in_file.add(id_interno)
        if ref_key:
            if ref_key in seen_refs_in_file:
                duplicate_in_file = True
            seen_refs_in_file.add(ref_key)

        if duplicate_in_file:
            report["duplicateInFile"] += 1
            mensagem = "Duplicidade detectada no proprio arquivo"
            row_details.append(
                {
                    "ordem": ordem,
                    "sheet_name": ctx.get("sheetName") or "XLSX",
                    "row_number": int(ctx.get("rowNumber") or 0),
                    "status_linha": "CONFLICT",
                    "id_interno": id_interno,
                    "ref_key": ref_key,
                    "mensagem": mensagem,
                }
            )
            normalized_source_rows.append(_serialize_preview_source_row(ctx, incoming, "CONFLICT", mensagem))
            continue

        by_id = existing_by_id.get(id_interno) if id_interno else None
        by_ref = existing_by_ref.get(ref_key) if ref_key else None

        if by_id:
            report["duplicateById"] += 1
        if not by_id and by_ref:
            report["duplicateByRef"] += 1

        if by_id and by_ref and by_id.id != by_ref.id:
            report["conflictIdVsRef"] += 1
            mensagem = "Conflito ID x chave referencia; revisar antes de aplicar"
            row_details.append(
                {
                    "ordem": ordem,
                    "sheet_name": ctx.get("sheetName") or "XLSX",
                    "row_number": int(ctx.get("rowNumber") or 0),
                    "status_linha": "CONFLICT",
                    "id_interno": id_interno,
                    "ref_key": ref_key,
                    "mensagem": mensagem,
                }
            )
            normalized_source_rows.append(_serialize_preview_source_row(ctx, incoming, "CONFLICT", mensagem))
            continue

        target = by_id or by_ref
        if not target:
            report["created"] += 1
            mensagem = "Registro novo pronto para entrar"
            row_details.append(
                {
                    "ordem": ordem,
                    "sheet_name": ctx.get("sheetName") or "XLSX",
                    "row_number": int(ctx.get("rowNumber") or 0),
                    "status_linha": "CREATED",
                    "id_interno": id_interno,
                    "ref_key": ref_key,
                    "mensagem": mensagem,
                }
            )
            normalized_source_rows.append(_serialize_preview_source_row(ctx, incoming, "CREATED", mensagem))
            if len(new_rows_preview) < 25:
                new_rows_preview.append(
                    {
                        "ordem": ordem,
                        "sheet_name": ctx.get("sheetName") or "XLSX",
                        "row_number": int(ctx.get("rowNumber") or 0),
                        "id_interno": id_interno,
                        "identificacao": incoming.get("identificacao") or "",
                        "deputado": incoming.get("deputado") or "",
                        "municipio": incoming.get("municipio") or "",
                        "cod_acao": incoming.get("cod_acao") or "",
                        "status_oficial": incoming.get("status_oficial") or "",
                        "valor_atual": _as_text(incoming.get("valor_atual")),
                        "ref_key": ref_key,
                        "mensagem": "Novo registro",
                    }
                )
            continue

        changed_fields = []
        for field_key, field_type in TRACKED_FIELDS:
            next_value = incoming.get(field_key)
            if not _has_incoming_value(next_value, field_type):
                continue
            current_value = getattr(target, field_key, None)
            if _field_changed(current_value, next_value, field_type):
                changed_fields.append(field_key)

        incoming_status = incoming.get("status_oficial")
        if incoming_status and _normalize_status(target.status_oficial) != _normalize_status(incoming_status):
            changed_fields.append("status_oficial")

        if changed_fields:
            report["updated"] += 1
            mensagem = f"Registro existente com alteracoes: {', '.join(changed_fields[:6])}"
            row_details.append(
                {
                    "ordem": ordem,
                    "sheet_name": ctx.get("sheetName") or "XLSX",
                    "row_number": int(ctx.get("rowNumber") or 0),
                    "status_linha": "UPDATED",
                    "id_interno": target.id_interno,
                    "ref_key": ref_key or _build_reference_key(
                        {
                            "identificacao": target.identificacao,
                            "cod_subfonte": target.cod_subfonte,
                            "cod_acao": target.cod_acao,
                            "municipio": target.municipio,
                            "deputado": target.deputado,
                        }
                    ),
                    "mensagem": mensagem,
                }
            )
            normalized_source_rows.append(_serialize_preview_source_row(ctx, incoming, "UPDATED", mensagem))
        else:
            report["unchanged"] += 1
            mensagem = "Registro existente sem alteracao"
            row_details.append(
                {
                    "ordem": ordem,
                    "sheet_name": ctx.get("sheetName") or "XLSX",
                    "row_number": int(ctx.get("rowNumber") or 0),
                    "status_linha": "UNCHANGED",
                    "id_interno": target.id_interno,
                    "ref_key": ref_key or _build_reference_key(
                        {
                            "identificacao": target.identificacao,
                            "cod_subfonte": target.cod_subfonte,
                            "cod_acao": target.cod_acao,
                            "municipio": target.municipio,
                            "deputado": target.deputado,
                        }
                    ),
                    "mensagem": mensagem,
                }
            )
            normalized_source_rows.append(_serialize_preview_source_row(ctx, incoming, "UNCHANGED", mensagem))

    report["sourceRows"] = _coerce_preview_source_rows(report.get("sourceRows") or [], row_details)
    return report


IMPORT_GOVERNANCE_ROLES = {"APG", "SUPERVISAO", "POWERBI", "PROGRAMADOR"}


def _can_govern_imports(actor: dict | None) -> bool:
    role = str((actor or {}).get("role") or "").strip().upper()
    return role in IMPORT_GOVERNANCE_ROLES


def _actor_matches_lote(lote: ImportLote, actor: dict | None) -> bool:
    if not lote or not actor:
        return False
    if _can_govern_imports(actor):
        return True
    actor_id = actor.get("id")
    if actor_id is not None and lote.usuario_id is not None:
        try:
            return int(actor_id) == int(lote.usuario_id)
        except Exception:
            pass
    actor_name = str(actor.get("name") or "").strip().lower()
    actor_role = str(actor.get("role") or "").strip().upper()
    return actor_name == str(lote.usuario_nome or "").strip().lower() and actor_role == str(lote.setor or "").strip().upper()


def _ensure_lote_access(lote: ImportLote | None, actor: dict | None) -> ImportLote:
    if not lote:
        raise HTTPException(status_code=404, detail="lote de importacao nao encontrado")
    if not _actor_matches_lote(lote, actor):
        raise HTTPException(status_code=403, detail="sem acesso a este lote")
    return lote


def _append_governance_log(
    *,
    lote_id: int,
    acao: str,
    motivo: str,
    actor: dict | None,
    detalhes: dict | None,
    db: Session,
    utcnow: Callable[[], datetime],
) -> None:
    db.add(
        ImportGovernancaLog(
            lote_id=lote_id,
            acao=str(acao or "").strip().upper(),
            motivo=str(motivo or "").strip(),
            usuario_id=(actor or {}).get("id"),
            usuario_nome=str((actor or {}).get("name") or ""),
            setor=str((actor or {}).get("role") or ""),
            detalhes_json=json.dumps(detalhes or {}, ensure_ascii=False),
            created_at=utcnow(),
        )
    )


def create_import_lot_service(
    *,
    payload,
    actor: dict,
    db: Session,
    resolve_event_origin: Callable[[str | None, dict | None, str], str],
    utcnow: Callable[[], datetime],
    broadcast_update: Callable[[str, int | None], None],
) -> dict:
    lote = ImportLote(
        arquivo_nome=payload.arquivo_nome,
        arquivo_hash=payload.arquivo_hash or "",
        linhas_lidas=max(0, int(payload.linhas_lidas or 0)),
        linhas_validas=max(0, int(payload.linhas_validas or 0)),
        linhas_ignoradas=max(0, int(payload.linhas_ignoradas or 0)),
        registros_criados=max(0, int(payload.registros_criados or 0)),
        registros_atualizados=max(0, int(payload.registros_atualizados or 0)),
        sem_alteracao=max(0, int(payload.sem_alteracao or 0)),
        duplicidade_id=max(0, int(payload.duplicidade_id or 0)),
        duplicidade_ref=max(0, int(payload.duplicidade_ref or 0)),
        duplicidade_arquivo=max(0, int(payload.duplicidade_arquivo or 0)),
        conflito_id_ref=max(0, int(payload.conflito_id_ref or 0)),
        abas_lidas=" | ".join([x for x in (payload.abas_lidas or []) if str(x).strip()]),
        observacao=payload.observacao or "",
        origem_evento=resolve_event_origin(payload.origem_evento, actor, "IMPORT"),
        usuario_id=actor.get("id"),
        usuario_nome=actor.get("name") or "",
        setor=actor.get("role") or "",
        status_governanca="APLICADO",
        governanca_motivo="",
        governado_por_id=None,
        governado_por_nome="",
        governado_por_setor="",
        governado_em=None,
        registros_removidos=0,
        created_at=utcnow(),
    )
    db.add(lote)
    db.flush()
    _append_governance_log(
        lote_id=lote.id,
        acao="CRIADO",
        motivo="Import aplicado na base.",
        actor=actor,
        detalhes={
            "arquivo_nome": lote.arquivo_nome,
            "linhas_lidas": lote.linhas_lidas,
            "linhas_ignoradas": lote.linhas_ignoradas,
            "registros_criados": lote.registros_criados,
            "registros_atualizados": lote.registros_atualizados,
        },
        db=db,
        utcnow=utcnow,
    )
    db.commit()
    db.refresh(lote)
    broadcast_update("import_lote", lote.id)
    return {"ok": True, "id": lote.id}


def list_import_lots_service(*, limit: int, actor: dict, db: Session) -> list[ImportLote]:
    query = db.query(ImportLote)
    if not _can_govern_imports(actor):
        actor_id = actor.get("id")
        if actor_id is not None:
            query = query.filter(ImportLote.usuario_id == actor_id)
        else:
            query = query.filter(
                ImportLote.usuario_nome == str(actor.get("name") or ""),
                ImportLote.setor == str(actor.get("role") or ""),
            )
    return query.order_by(ImportLote.created_at.desc(), ImportLote.id.desc()).limit(limit).all()


def create_import_lines_service(
    *,
    payload,
    actor: dict,
    db: Session,
    utcnow: Callable[[], datetime],
    broadcast_update: Callable[[str, int | None], None],
) -> dict:
    lote = _ensure_lote_access(db.get(ImportLote, payload.lote_id), actor)

    linhas = payload.linhas or []
    if not linhas:
        return {"ok": True, "inserted": 0}

    inserted = 0
    now = utcnow()
    for ln in linhas:
        db.add(
            ImportLinha(
                lote_id=lote.id,
                ordem=max(0, int(ln.ordem or 0)),
                sheet_name=(ln.sheet_name or "")[:120],
                row_number=max(0, int(ln.row_number or 0)),
                status_linha=(ln.status_linha or "UNCHANGED").upper(),
                id_interno=(ln.id_interno or "")[:60],
                ref_key=(ln.ref_key or "")[:255],
                mensagem=ln.mensagem or "",
                created_at=now,
            )
        )
        inserted += 1

    db.commit()
    broadcast_update("import_linha", lote.id)
    return {"ok": True, "inserted": inserted}


def list_import_lines_service(*, lote_id: int, limit: int, actor: dict, db: Session) -> list[ImportLinha]:
    lote = _ensure_lote_access(db.get(ImportLote, lote_id), actor)
    return (
        db.query(ImportLinha)
        .filter(ImportLinha.lote_id == lote.id)
        .order_by(ImportLinha.ordem.asc(), ImportLinha.id.asc())
        .limit(limit)
        .all()
    )


def list_import_governance_logs_service(*, lote_id: int, limit: int, actor: dict, db: Session) -> list[ImportGovernancaLog]:
    lote = _ensure_lote_access(db.get(ImportLote, lote_id), actor)
    return (
        db.query(ImportGovernancaLog)
        .filter(ImportGovernancaLog.lote_id == lote.id)
        .order_by(ImportGovernancaLog.created_at.desc(), ImportGovernancaLog.id.desc())
        .limit(limit)
        .all()
    )


def govern_import_lot_service(
    *,
    lote_id: int,
    payload,
    actor: dict,
    db: Session,
    utcnow: Callable[[], datetime],
    broadcast_update: Callable[[str, int | None], None],
) -> ImportLote:
    lote = db.get(ImportLote, lote_id)
    if not lote:
        raise HTTPException(status_code=404, detail="lote de importacao nao encontrado")
    if not _can_govern_imports(actor):
        raise HTTPException(status_code=403, detail="perfil sem permissao para governar imports")

    acao = str(payload.acao or "").strip().upper()
    motivo = str(payload.motivo or "").strip()
    now = utcnow()

    if acao == "CORRIGIR":
        lote.status_governanca = "CORRIGIDO"
        lote.governanca_motivo = motivo
        lote.governado_por_id = actor.get("id")
        lote.governado_por_nome = str(actor.get("name") or "")
        lote.governado_por_setor = str(actor.get("role") or "")
        lote.governado_em = now
        _append_governance_log(
            lote_id=lote.id,
            acao="CORRIGIR",
            motivo=motivo,
            actor=actor,
            detalhes={"status_governanca": "CORRIGIDO"},
            db=db,
            utcnow=utcnow,
        )
        db.commit()
        db.refresh(lote)
        broadcast_update("import_lote", lote.id)
        return lote

    if acao != "REMOVER":
        raise HTTPException(status_code=400, detail="acao de governanca invalida")

    import_lines = (
        db.query(ImportLinha)
        .filter(ImportLinha.lote_id == lote.id, ImportLinha.status_linha == "CREATED")
        .order_by(ImportLinha.id.asc())
        .all()
    )
    created_ids = []
    for line in import_lines:
        raw_id = str(line.id_interno or "").strip()
        if raw_id:
            created_ids.append(raw_id)
    created_ids = list(dict.fromkeys(created_ids))

    emendas = []
    if created_ids:
        emendas = db.query(Emenda).filter(Emenda.id_interno.in_(created_ids)).all()

    removable = []
    blocked = []
    for emenda in emendas:
        has_children = db.query(Emenda.id).filter(Emenda.parent_id == emenda.id).first() is not None
        if has_children:
            blocked.append(emenda.id_interno)
            continue
        removable.append(emenda)

    removable_ids = [int(item.id) for item in removable]
    removable_backend_ids = [item.id_interno for item in removable]

    if removable_ids:
        db.query(SupportThread).filter(SupportThread.emenda_id.in_(removable_ids)).update(
            {"emenda_id": None},
            synchronize_session=False,
        )
        db.query(EmendaLock).filter(EmendaLock.emenda_id.in_(removable_ids)).delete(synchronize_session=False)
        db.query(Historico).filter(Historico.emenda_id.in_(removable_ids)).delete(synchronize_session=False)
        for emenda in removable:
            db.delete(emenda)

    lote.status_governanca = "REMOVIDO"
    lote.governanca_motivo = motivo
    lote.governado_por_id = actor.get("id")
    lote.governado_por_nome = str(actor.get("name") or "")
    lote.governado_por_setor = str(actor.get("role") or "")
    lote.governado_em = now
    lote.registros_removidos = len(removable_backend_ids)

    _append_governance_log(
        lote_id=lote.id,
        acao="REMOVER",
        motivo=motivo,
        actor=actor,
        detalhes={
            "registros_removidos": removable_backend_ids,
            "registros_bloqueados": blocked,
            "linhas_created_lote": created_ids,
        },
        db=db,
        utcnow=utcnow,
    )

    db.commit()
    db.refresh(lote)
    broadcast_update("import_lote", lote.id)
    if removable_ids:
        broadcast_update("emenda", None)
    return lote


def create_export_log_service(
    *,
    payload,
    actor: dict,
    db: Session,
    resolve_event_origin: Callable[[str | None, dict | None, str], str],
    utcnow: Callable[[], datetime],
    broadcast_update: Callable[[str, int | None], None],
) -> dict:
    log = ExportLog(
        formato=payload.formato,
        arquivo_nome=payload.arquivo_nome,
        quantidade_registros=max(0, int(payload.quantidade_registros or 0)),
        quantidade_eventos=max(0, int(payload.quantidade_eventos or 0)),
        filtros_json=payload.filtros_json or "",
        modo_headers=(payload.modo_headers or "normalizados")[:30],
        escopo_exportacao=(payload.escopo_exportacao or "ATUAIS")[:20],
        round_trip_ok=payload.round_trip_ok,
        round_trip_issues=" | ".join([x for x in (payload.round_trip_issues or []) if str(x).strip()]),
        origem_evento=resolve_event_origin(payload.origem_evento, actor, "EXPORT"),
        usuario_id=actor.get("id"),
        usuario_nome=actor.get("name") or "",
        setor=actor.get("role") or "",
        created_at=utcnow(),
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    broadcast_update("export_log", log.id)
    return {"ok": True, "id": log.id}


def list_export_logs_service(*, limit: int, db: Session) -> list[ExportLog]:
    return db.query(ExportLog).order_by(ExportLog.created_at.desc(), ExportLog.id.desc()).limit(limit).all()
